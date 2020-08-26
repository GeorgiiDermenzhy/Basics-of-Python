"""Task 4 of Module 4."""
from tkinter.filedialog import askopenfilename
from pathlib import Path
import os
from openpyxl import load_workbook


def spreadsheet_cell_inventer():
    """Transpose values in workbook."""
    file = askopenfilename(title='Select excel file')
    file_path = Path(file)
    os.chdir(file_path.parent)

    wb = load_workbook(file_path)
    ws = wb.active

    new_ws = wb.create_sheet("Transposed data")

    for column in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            new_ws.cell(column, row).value = ws.cell(row, column).value

    wb.save(file_path.name)


spreadsheet_cell_inventer()
